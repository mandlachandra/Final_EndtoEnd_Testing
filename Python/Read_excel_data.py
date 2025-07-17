from openpyxl import load_workbook

#load the excel workbook
workbook = load_workbook("data.xlsx")

#select the sheet by name
sheet = workbook["sheet1"]

#Get the value of a specific cell(row=2,column=1)
username = sheet.cell(row=2,column=1).value
password = sheet.cell(row=2,column=2).value

print(f"username: {username},password:{password}")

#loop through the cells
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=2):
    uname = row[0].value
    pwd = row[1].value
    print(f"username:{uname},password:{pwd}")