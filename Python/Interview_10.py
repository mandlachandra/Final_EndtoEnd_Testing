from openpyxl import load_workbook

fruits1 = ["apple", "banana", "cherry", "Mango", "Orange"]
fruits2 = ["apple", "banana", "cherry", "palm", "guava", "pier"]

s1 = set(fruits1)
s2 = set(fruits2)
common_fruits = s1 & s2
print("common_fruits",common_fruits)

workbook = load_workbook("C://Users//DELL//OneDrive//Desktop//New Microsoft Excel Worksheet (2).xlsx")
sheet = workbook.active
sheet.title = "common_fruits_names"

sheet.cell(row=1,column=1,value="fruit name")

for idx, fruit in enumerate(common_fruits,start=2):
    sheet.cell(row=idx,column=1,value=fruit)

workbook.save("common_fruit_names")

class Person:

    def __init__(self,name,age,salary):
        self.name = name
        self.age = age
        self.salary = salary

    # def __str__(self):
    #     return f"name:{self.name}, Age:{self.age},salary:{self.salary}"

    def __str__(self):
        return f"name:{self.name},Age:{self.age},salary:{self.salary}"

    #creata a list of person objects
people = [
    Person("Alice",30,50000),
    Person("bob",28,49000),
    Person("charlie",27,45000),
    Person("max",25,30000)
    ]

#sort by age
people_sorted = sorted(people, key=lambda people: people.age)
#print the sorted list
print("Sorted by age:")
for people in people_sorted:
    # print(people.age, people.name, people.salary)
    print(people)








