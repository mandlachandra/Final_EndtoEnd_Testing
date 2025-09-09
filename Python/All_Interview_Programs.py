
def flatten_list(nested_list):
    flat =[]

    for item in nested_list:
        if isinstance(item,list):
            flat.extend(flatten_list(item))
        else:
            flat.append(item)

    return flat
input_data = [3, 5, [9, [7, 8, [12, 71]]]]
output = flatten_list(input_data)
print("The output is:",output)

#2nd most character
str = "secondmostfrequentcharacter"
from collections import Counter
freq = Counter(str)
sorted_freq = freq.most_common()

if len(sorted_freq)>=2:
    second_most_character, second_most_count = sorted_freq[1]
    print(f"2nd most character is {second_most_character} with count {second_most_count}")
else:
    print("invalid count")

#3 Shallow copy = creates new creates a new object but does not create copies of nested objects
import copy

original = [[1,2,3],[4,5,6]]
shallow = copy.copy(original)

original[0][0] =99
print(shallow[0][0])

#deep copy = deep copy will create completely independent object including all nested objects
import copy
origianl = [[1,2],[3,4]]
deep = copy.deepcopy(original)

original[0][0] =99
print(deep[0][0])

#4th question
fruits1 = ["apple", "banana", "cherry", "Mango", "Orange"]
fruits2 = ["apple", "banana", "cherry", "palm", "guava", "pier"]

s1 = set(fruits1)
s2= set(fruits2)
common_fruits = s1 & s2
print("common_fruits",common_fruits)

from openpyxl import load_workbook
workbook = load_workbook("C://Users//DELL//OneDrive//Desktop//New Microsoft Excel Worksheet (2).xlsx")
sheet = workbook.active
sheet_title = "common_fruit_names"

sheet.cell(row=1,column=1,value="fruit_name")

for idx, fruit in enumerate("common_fruits",start=2):
    sheet.cell(row=idx,column=1,value=fruit)

workbook.save("common_fruit_names2")

##5th.count frequency of each character in a string
str = "mandlachandrasekhar"
freq = {}

for char in str:
    if char in freq:
        freq[char]+=1

    else:
        freq[char] =1

print(freq)
#6th pattern problem
data = {"a": 1, "b": 2, "c": 3}

forward = ''.join(k * v for k,v in data.items())
backward = ''.join(k *v for k , v in reversed(data.items()))
output = forward + "-" + backward
print(output)
#output abbccc-cccbba

#7th read data from excel
# from openpyxl import load_workbook
#
# workbook = load_workbook("data.xlsx")
# sheet = workbook["Sheet1"]
#
# username = sheet.cell(row=2,column=1).value
# password = sheet.cell(row=2,column=2).value
#
# print(f"username is {username} and passwod is {password}")
#
# #through loop
# for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=2):
#     uname = row[0].value
#     pwd = row[1].value
#
#     print(f"username is {username} and password is {password}")

#8th Multi Threading
import threading
import time

# def print_numbers():
#     for i in range(5):
#         print("Thread",i)
#         time.sleep(1)
#
# #create 2 threads
# t1 = threading.Thread(target=print_numbers)
# t2 = threading.Thread(target=print_numbers)
#
# t1.start()
# t2.start()
#
# t1.join()
# t2.join()
#
# print("done")

#multi processing
# import multiprocessing
# import time
#
# def print_numbers():
#     for i in range(5):
#         print("Thread",i)
#         time.sleep(1)
#
# t1=multiprocessing.Process(target=print_numbers)
# t2=multiprocessing.Process(target=print_numbers)
#
# t1.start()
# t2.start()
#
# t1.join()
# t2.join()
# print("done")

# class BankAccount:
#
#     def __init__(self,account_number,pin,balance=0):
#         self.account_number = account_number
#         self.__pin = pin
#         self.balance = balance
#
#     def check_balance(self,pin):
#         if pin == self.__pin:
#             return self.balance
#         else:
#             raise ValueError("Invalid pin")
#
#     def deposit(self,pin,amount):
#         if pin!=self.__pin:
#             return ValueError("Invalid pin Transaction denied")
#         if amount<=0:
#             return ValueError("Amount should be greater than 0")
#         self.balance+=amount
#         return self.balance
#
#     def update_pin(self,old_pin,new_pin):
#         if old_pin!= self.__pin:
#             raise ValueError("Invalid pin")
#         self.__pin = new_pin
#         return True
#
# account = BankAccount(account_number=123456,pin=1234,balance=10000000)

def divide(a, b):
    try:

        result = a / b
        print("result", result)
    except ZeroDivisionError:
        print("division by zero is not allowd")

    finally:
        print("execution is done")


divide(10, 2)
divide(10, 0)

# import pytest
# def test_capture_exception():
#     with pytest.raises(ZeroDivisionError) as excinfo:
#         divide(10,0)
#     assert str(excinfo.value) == "division by zero"

#Failed screenshot
# @pytest.hookimpl(tryfirst=True, hookwrapper=True)
# def pytest_runtest_makereport(item, call):
#     outcome = yield
#
#     rep = outcome.get_result():
#     if rep.when == "call" and rep.failed:
#         driver = item.funcargs['driver']
#         driver.save_screenshot(f"screenshot/{item.name}.png")
#         print(f"\n screenshot for failed test :{item.name}.png")













