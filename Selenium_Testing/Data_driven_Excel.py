import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

def get_data():
    workbook = openpyxl.load_workbook("testdata.xlsx")
    sheet = workbook.active
    rows = sheet.max_row
    data = []

    for i in range(2,rows+1):
        username = sheet.cell(i,1).value
        password = sheet.cell(i,2).value
        data.append((username,password))
    return data

@pytest.mark.parametrize("username,password",get_data())
def test_login(username,password):
    driver=webdriver.Chrome()
    driver.get("https://www.google.com")
    driver.find_element(By.ID,"username").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"loginbutton").click()
    driver.quit()