# from selenium import webdriver
# from openpyxl import load_workbook
# from selenium.webdriver.common.by import By
# import time
#
#
# driver = webdriver.Chrome()
# driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
# driver.maximize_window()
# time.sleep(2)
#
# #Read data from excel
# wb = load_workbook("C://Users//DELL//OneDrive//Desktop//Credentials.xlsx")
# sheet = wb["Sheet1"]
#
# for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=2):
#     username = row[0].value
#     password = row[1].value
#
#     #Enter username and password in the login form
#     driver.find_element(By.NAME,"username").send_keys(username)
#     driver.find_element(By.NAME,"password").send_keys(password)
#     driver.find_element(By.XPATH, "//button[@type='submit']").click()
#
#     print("success")
#
#
#     driver.back()

#create a utility function to read data
from openpyxl import load_workbook

def get_excel_data(file_path,sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    data = []

    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(row) #(username , password)
    return data

#Pytest test file with Parameterization

import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
#from utils import get_excel_data
import time

#read data from excel
test_data = get_excel_data("C://Users//DELL//Desktop//Credentials.xlsx", "Sheet1")

@pytest.mark.parametrize("username,password",test_data)
def test_login(username,password):
    driver=webdriver.Chrome()
    driver.maximize_window()
    time.sleep(4)

    driver.find_element(By.NAME,"username").send_keys(username)
    driver.find_element(By.NAME,"password").send_keys(password)
    driver.find_element(By.XPATH,"//button[@type='submit']").click()
    time.sleep(5)

    if "dashboard" in driver.current_url.lower():
        print(f"login successful for :{username}")
    else:
        print(f"login failed for: {username}")

    driver.quit()


